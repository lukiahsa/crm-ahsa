"""Parser dan transaction engine untuk import master customer XLSX."""

from __future__ import annotations

import hashlib
import io
import json
import re
import sqlite3
import unicodedata
from collections import OrderedDict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import BinaryIO, Iterable

from openpyxl import load_workbook


MAX_IMPORT_BYTES = 10 * 1024 * 1024
MAX_IMPORT_ROWS = 20_000

STATUS_PROSPECT = "Prospek"
STATUS_EXISTING = "Existing Customer"
UNCLASSIFIED = "Belum Terklasifikasi"


@dataclass(frozen=True)
class SheetDefinition:
    headers: tuple[str, ...]
    status: str
    interest: str | None = None


SHEET_DEFINITIONS: OrderedDict[str, SheetDefinition] = OrderedDict(
    (
        (
            "Customer Tempat Sampah",
            SheetDefinition(("Nama", "Nomor WhatsApp"), STATUS_PROSPECT, "Tempat Sampah"),
        ),
        (
            "Customer Tangga",
            SheetDefinition(("Nama", "Nomor WhatsApp"), STATUS_PROSPECT, "Tangga"),
        ),
        (
            "Customer MH",
            SheetDefinition(("Nama", "Nomor WhatsApp"), STATUS_PROSPECT, "Material Handling"),
        ),
        (
            "Belum Terklasifikasi",
            SheetDefinition(("Nama", "Nomor WhatsApp"), STATUS_PROSPECT),
        ),
        (
            "Existing Produk Tempat Sampah",
            SheetDefinition(("Nama", "Nomor WhatsApp"), STATUS_EXISTING, "Tempat Sampah"),
        ),
        (
            "Existing Produk Tangga",
            SheetDefinition(("Nama", "Nomor WhatsApp"), STATUS_EXISTING, "Tangga"),
        ),
        (
            "Existing Produk MH",
            SheetDefinition(
                (
                    "No",
                    "Nama PIC",
                    "Kontak WA",
                    "Perusahaan",
                    "Produk Dibeli",
                    "Email",
                    "Alamat",
                ),
                STATUS_EXISTING,
            ),
        ),
    )
)

PRODUCT_KEYWORDS: OrderedDict[str, tuple[str, ...]] = OrderedDict(
    (
        ("Tempat Sampah", ("dustbin", "dust", "trash", "sampah")),
        ("Tangga", ("tangga", "ladder", "telescopic", "fiberglass")),
        (
            "Material Handling",
            (
                "hand pallet",
                "stacker",
                "forklift",
                "scissor",
                "lift table",
                "reach truck",
                "hand truck",
                "pallet mover",
                "drum",
            ),
        ),
    )
)

EMAIL_PATTERN = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")
PHONE_PUNCTUATION = re.compile(r"[\s+\-()]", re.UNICODE)
MULTI_VALUE_SEPARATOR = "; "


class CustomerImportError(ValueError):
    """Error validasi fatal yang harus menghentikan seluruh import."""


def clean_text(value) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text or None


def _text_from_excel(value) -> tuple[str | None, list[str]]:
    warnings: list[str] = []
    if value is None:
        return None, warnings

    if isinstance(value, bool):
        return str(value), ["Nomor WhatsApp bertipe boolean dan tidak valid."]

    if isinstance(value, int):
        warnings.append("Nomor WhatsApp tersimpan sebagai angka Excel; verifikasi digit sumber.")
        return str(value), warnings

    if isinstance(value, float):
        warnings.append("Nomor WhatsApp tersimpan sebagai angka Excel; verifikasi digit sumber.")
        if not value.is_integer():
            return str(value), warnings
        return format(Decimal(str(value)), "f").split(".", 1)[0], warnings

    text = clean_text(value)
    if text and re.fullmatch(r"[+-]?\d+(?:\.\d+)?[Ee][+-]?\d+", text):
        warnings.append("Nomor WhatsApp memakai notasi ilmiah; verifikasi digit sumber.")
        try:
            number = Decimal(text)
        except InvalidOperation:
            return text, warnings
        if number == number.to_integral_value():
            return format(number, "f").split(".", 1)[0], warnings
    return text, warnings


def normalize_whatsapp(value) -> dict:
    raw, warnings = _text_from_excel(value)
    if not raw:
        return {
            "raw": None,
            "normalized": None,
            "valid": False,
            "missing": True,
            "warnings": warnings,
        }

    digits = PHONE_PUNCTUATION.sub("", raw)
    if not digits.isdigit():
        warnings.append("Nomor WhatsApp mengandung karakter selain digit yang diizinkan.")
        return {
            "raw": raw,
            "normalized": None,
            "valid": False,
            "missing": False,
            "warnings": warnings,
        }

    if digits.startswith("0"):
        digits = "62" + digits[1:]
    elif not digits.startswith("62"):
        warnings.append("Nomor WhatsApp tidak diawali 0 atau 62.")
        return {
            "raw": raw,
            "normalized": None,
            "valid": False,
            "missing": False,
            "warnings": warnings,
        }

    if not digits.startswith("628"):
        warnings.append("Nomor bukan nomor seluler Indonesia berawalan 628.")
    elif not 10 <= len(digits) <= 15:
        warnings.append("Panjang nomor WhatsApp di luar rentang 10–15 digit.")
    else:
        return {
            "raw": raw,
            "normalized": digits,
            "valid": True,
            "missing": False,
            "warnings": warnings,
        }

    return {
        "raw": raw,
        "normalized": None,
        "valid": False,
        "missing": False,
        "warnings": warnings,
    }


def normalize_email(value) -> dict:
    raw = clean_text(value)
    if not raw:
        return {"raw": None, "normalized": None, "valid": False, "missing": True}
    normalized = raw.casefold()
    return {
        "raw": raw,
        "normalized": normalized if EMAIL_PATTERN.fullmatch(normalized) else None,
        "valid": bool(EMAIL_PATTERN.fullmatch(normalized)),
        "missing": False,
    }


def classify_product_keywords(value) -> list[str]:
    text = clean_text(value)
    if not text:
        return []
    folded = text.casefold()
    matches: list[str] = []
    for category, keywords in PRODUCT_KEYWORDS.items():
        if any(re.search(rf"(?<!\w){re.escape(keyword)}(?!\w)", folded) for keyword in keywords):
            matches.append(category)
    return matches


def normalize_name(value) -> dict:
    original = clean_text(value)
    if not original:
        return {"original": None, "normalized": None}

    searchable = "".join(
        character
        for character in original
        if unicodedata.category(character)[0] in {"L", "N", "M", "P", "Z"}
    )
    searchable = re.sub(r"\s+", " ", searchable).strip()
    for keywords in PRODUCT_KEYWORDS.values():
        for keyword in sorted(keywords, key=len, reverse=True):
            searchable = re.sub(
                rf"(?<!\w){re.escape(keyword)}(?!\w)",
                " ",
                searchable,
                flags=re.IGNORECASE,
            )
    normalized = re.sub(r"\s+", " ", searchable).strip(" .,/|_-()").strip()
    return {
        "original": original,
        "normalized": normalized or original,
    }


def _normalized_header(value) -> str:
    return clean_text(value) or ""


def _row_is_empty(values: Iterable) -> bool:
    return all(clean_text(value) is None for value in values)


def _build_record(sheet_name: str, row_number: int, row: dict) -> dict:
    is_existing_mh = sheet_name == "Existing Produk MH"
    source_name = row.get("Nama PIC") if is_existing_mh else row.get("Nama")
    source_phone = row.get("Kontak WA") if is_existing_mh else row.get("Nomor WhatsApp")
    name = normalize_name(source_name)
    phone = normalize_whatsapp(source_phone)
    email = normalize_email(row.get("Email"))
    definition = SHEET_DEFINITIONS[sheet_name]

    warnings = list(phone["warnings"])
    errors: list[str] = []
    if not email["missing"] and not email["valid"]:
        warnings.append("Format email tidak valid; email hanya dipertahankan sebagai data sumber.")
    if phone["missing"]:
        warnings.append("Nomor WhatsApp kosong.")
    elif not phone["valid"]:
        warnings.append("Nomor WhatsApp tidak akan disimpan sebagai nomor valid.")
    if not name["original"] and phone["valid"]:
        warnings.append("Nama kosong; nomor WhatsApp valid digunakan sebagai identifier tampilan.")
    elif not name["original"]:
        warnings.append("Nama kosong.")
    if not name["original"] and not phone["valid"]:
        errors.append("Nama atau nomor WhatsApp valid wajib tersedia minimal salah satu.")

    keyword_source = row.get("Produk Dibeli") if is_existing_mh else source_name
    keyword_products = classify_product_keywords(keyword_source)
    if definition.interest:
        interests = [definition.interest]
    elif len(keyword_products) == 1:
        interests = keyword_products
    else:
        interests = []
        if len(keyword_products) > 1:
            warnings.append("Keyword produk ambigu; klasifikasi sumber dipertahankan.")

    return {
        "source_sheet": sheet_name,
        "source_row": row_number,
        "nama_asli": name["original"],
        "nama_normalisasi": name["normalized"],
        "whatsapp_raw": phone["raw"],
        "whatsapp_normalized": phone["normalized"],
        "email_raw": email["raw"],
        "email": email["normalized"],
        "instansi": clean_text(row.get("Perusahaan")),
        "alamat": clean_text(row.get("Alamat")),
        "produk_existing": clean_text(row.get("Produk Dibeli")),
        "interests": interests,
        "keyword_products": keyword_products,
        "sumber": sheet_name,
        "status": definition.status,
        "klasifikasi_produk": (
            MULTI_VALUE_SEPARATOR.join(interests) if interests else UNCLASSIFIED
        ),
        "warnings": warnings,
        "errors": errors,
    }


def parse_customer_workbook(source: bytes | BinaryIO) -> dict:
    if isinstance(source, bytes):
        stream = io.BytesIO(source)
    else:
        stream = source

    try:
        workbook = load_workbook(
            stream,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:  # openpyxl raises several format-specific errors
        raise CustomerImportError(f"File XLSX tidak dapat dibaca: {exc}") from exc

    expected_sheets = list(SHEET_DEFINITIONS)
    if workbook.sheetnames != expected_sheets:
        workbook.close()
        missing = [name for name in expected_sheets if name not in workbook.sheetnames]
        extra = [name for name in workbook.sheetnames if name not in expected_sheets]
        raise CustomerImportError(
            "Struktur sheet tidak sesuai sumber resmi. "
            f"Sheet hilang: {missing or '-'}; sheet tambahan: {extra or '-'}; "
            f"urutan diterima: {workbook.sheetnames}."
        )

    records: list[dict] = []
    per_sheet: OrderedDict[str, dict] = OrderedDict()
    try:
        for sheet_name, definition in SHEET_DEFINITIONS.items():
            worksheet = workbook[sheet_name]
            rows = worksheet.iter_rows(values_only=True)
            try:
                header_values = next(rows)
            except StopIteration as exc:
                raise CustomerImportError(f"Sheet {sheet_name} tidak memiliki header.") from exc

            actual_headers = tuple(_normalized_header(value) for value in header_values)
            while actual_headers and actual_headers[-1] == "":
                actual_headers = actual_headers[:-1]
            if actual_headers != definition.headers:
                raise CustomerImportError(
                    f"Header sheet {sheet_name} tidak sesuai. "
                    f"Diharapkan {definition.headers}, ditemukan {actual_headers}."
                )

            sheet_records: list[dict] = []
            for row_number, values in enumerate(rows, start=2):
                relevant_values = values[: len(definition.headers)]
                if _row_is_empty(relevant_values):
                    continue
                if len(records) >= MAX_IMPORT_ROWS:
                    raise CustomerImportError(
                        f"Jumlah baris melebihi batas aman {MAX_IMPORT_ROWS}."
                    )
                row = dict(zip(definition.headers, relevant_values))
                record = _build_record(sheet_name, row_number, row)
                records.append(record)
                sheet_records.append(record)

            per_sheet[sheet_name] = {
                "rows": len(sheet_records),
                "valid_phone": sum(
                    bool(record["whatsapp_normalized"]) for record in sheet_records
                ),
                "warning_rows": sum(bool(record["warnings"]) for record in sheet_records),
                "error_rows": sum(bool(record["errors"]) for record in sheet_records),
                "status": definition.status,
            }
    finally:
        workbook.close()

    return {
        "records": records,
        "per_sheet": per_sheet,
        "total_rows": len(records),
    }


class _UnionFind:
    def __init__(self, size: int):
        self.parent = list(range(size))

    def find(self, item: int) -> int:
        while self.parent[item] != item:
            self.parent[item] = self.parent[self.parent[item]]
            item = self.parent[item]
        return item

    def union(self, first: int, second: int) -> None:
        first_root = self.find(first)
        second_root = self.find(second)
        if first_root != second_root:
            self.parent[second_root] = first_root


def _identity_keys(record: dict) -> list[tuple]:
    keys: list[tuple] = []
    if record.get("whatsapp_normalized"):
        keys.append(("whatsapp", record["whatsapp_normalized"]))
    if record.get("email"):
        keys.append(("email", record["email"].casefold()))
    name = clean_text(record.get("nama_normalisasi"))
    company = clean_text(record.get("instansi"))
    if name and company:
        keys.append(("name_company", name.casefold(), company.casefold()))
    return keys


def _ordered_unique(values: Iterable[str | None]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = clean_text(value)
        if not text:
            continue
        key = text.casefold()
        if key not in seen:
            seen.add(key)
            result.append(text)
    return result


def _join_unique(values: Iterable[str | None]) -> str | None:
    result = _ordered_unique(values)
    return MULTI_VALUE_SEPARATOR.join(result) if result else None


def _most_complete(values: Iterable[str | None]) -> str | None:
    candidates = _ordered_unique(values)
    if not candidates:
        return None
    return max(enumerate(candidates), key=lambda item: (len(item[1]), -item[0]))[1]


def _aggregate_group(group: list[dict]) -> dict:
    sources = _ordered_unique(record["sumber"] for record in group)
    interests = _ordered_unique(
        interest for record in group for interest in record["interests"]
    )
    names = _ordered_unique(record["nama_asli"] for record in group)
    normalized_names = _ordered_unique(record["nama_normalisasi"] for record in group)
    phones = _ordered_unique(record["whatsapp_normalized"] for record in group)
    emails = _ordered_unique(record["email"] for record in group)
    companies = _ordered_unique(record["instansi"] for record in group)
    addresses = _ordered_unique(record["alamat"] for record in group)

    conflicts: list[str] = []
    for label, values in (
        ("nomor WhatsApp", phones),
        ("email", emails),
        ("perusahaan", companies),
        ("alamat", addresses),
    ):
        if len(values) > 1:
            conflicts.append(f"Konflik {label}: {MULTI_VALUE_SEPARATOR.join(values)}")

    if len(group) > 1:
        conflicts.append(
            f"{len(group)} baris sumber digabung berdasarkan identifier duplicate."
        )

    phone = phones[0] if phones else None
    display_name = _most_complete(names) or phone
    warnings = _ordered_unique(
        warning for record in group for warning in record["warnings"]
    )
    if not names and phone:
        warnings.append("Nama kosong; identifier WhatsApp dipakai sebagai tampilan legacy.")

    source_sheets = _ordered_unique(record["source_sheet"] for record in group)
    return {
        "nama": display_name,
        "nama_asli": _join_unique(names),
        "nama_normalisasi": _most_complete(normalized_names) or phone,
        "whatsapp": phone,
        "whatsapp_raw": _join_unique(record["whatsapp_raw"] for record in group),
        "whatsapp_normalized": phone,
        "email_raw": _join_unique(record["email_raw"] for record in group),
        "email": emails[0] if emails else None,
        "instansi": _most_complete(companies),
        "alamat": _most_complete(addresses),
        "produk": _join_unique(interests),
        "produk_existing": _join_unique(record["produk_existing"] for record in group),
        "sumber": _join_unique(sources),
        "status": (
            STATUS_EXISTING
            if any(record["status"] == STATUS_EXISTING for record in group)
            else STATUS_PROSPECT
        ),
        "klasifikasi_produk": _join_unique(interests) or UNCLASSIFIED,
        "warnings": warnings,
        "errors": _ordered_unique(error for record in group for error in record["errors"]),
        "conflicts": conflicts,
        "source_sheets": source_sheets,
        "source_rows": [
            f"{record['source_sheet']}!{record['source_row']}" for record in group
        ],
        "source_row_count": len(group),
        "cross_sheet_duplicate": len(source_sheets) > 1,
    }


def consolidate_customer_records(parsed: dict) -> dict:
    valid_records = [record for record in parsed["records"] if not record["errors"]]
    error_records = [record for record in parsed["records"] if record["errors"]]
    union_find = _UnionFind(len(valid_records))
    key_owner: dict[tuple, int] = {}
    for index, record in enumerate(valid_records):
        for key in _identity_keys(record):
            owner = key_owner.setdefault(key, index)
            union_find.union(index, owner)

    grouped: OrderedDict[int, list[dict]] = OrderedDict()
    for index, record in enumerate(valid_records):
        grouped.setdefault(union_find.find(index), []).append(record)

    customers = [_aggregate_group(group) for group in grouped.values()]
    duplicate_in_file = sum(customer["source_row_count"] - 1 for customer in customers)
    duplicate_cross_sheet = sum(
        customer["source_row_count"] - 1
        for customer in customers
        if customer["cross_sheet_duplicate"]
    )
    summary = {
        "total_rows": parsed["total_rows"],
        "valid_phone": sum(
            bool(record["whatsapp_normalized"]) for record in parsed["records"]
        ),
        "warning_rows": sum(bool(record["warnings"]) for record in parsed["records"]),
        "error_rows": len(error_records),
        "duplicate_in_file": duplicate_in_file,
        "duplicate_cross_sheet": duplicate_cross_sheet,
        "unique_ready": len(customers),
        "existing_customers": sum(
            customer["status"] == STATUS_EXISTING for customer in customers
        ),
        "prospects": sum(customer["status"] == STATUS_PROSPECT for customer in customers),
        "unclassified": sum(
            customer["klasifikasi_produk"] == UNCLASSIFIED for customer in customers
        ),
    }
    return {
        "customers": customers,
        "error_records": error_records,
        "per_sheet": parsed["per_sheet"],
        "summary": summary,
    }


def _split_multi_value(value: str | None) -> list[str]:
    text = clean_text(value)
    if not text:
        return []
    return _ordered_unique(part for part in text.split(";") if clean_text(part))


def _merge_multi_value(existing: str | None, imported: str | None) -> str | None:
    return _join_unique([*_split_multi_value(existing), *_split_multi_value(imported)])


def _row_value(row: sqlite3.Row | dict, key: str):
    try:
        return row[key]
    except (KeyError, IndexError):
        return None


def _database_identity_keys(row: sqlite3.Row | dict) -> list[tuple]:
    phone = normalize_whatsapp(
        _row_value(row, "whatsapp_normalized") or _row_value(row, "whatsapp")
    )
    email = normalize_email(_row_value(row, "email"))
    normalized_name = (
        clean_text(_row_value(row, "nama_normalisasi"))
        or normalize_name(_row_value(row, "nama"))["normalized"]
    )
    company = clean_text(_row_value(row, "instansi"))
    keys: list[tuple] = []
    if phone["valid"]:
        keys.append(("whatsapp", phone["normalized"]))
    if email["valid"]:
        keys.append(("email", email["normalized"]))
    if normalized_name and company:
        keys.append(("name_company", normalized_name.casefold(), company.casefold()))
    return keys


def _find_database_matches(conn: sqlite3.Connection) -> dict[tuple, sqlite3.Row]:
    matches: dict[tuple, sqlite3.Row] = {}
    for row in conn.execute("SELECT * FROM customers ORDER BY id").fetchall():
        for key in _database_identity_keys(row):
            matches.setdefault(key, row)
    return matches


def _merge_status(existing: str | None, imported: str) -> str:
    existing = clean_text(existing)
    if not existing:
        return imported
    if imported == STATUS_EXISTING and existing in {
        STATUS_PROSPECT,
        "Follow Up",
        "Penawaran",
    }:
        return STATUS_EXISTING
    return existing


def _build_non_destructive_updates(existing: sqlite3.Row, imported: dict) -> dict:
    updates: dict = {}
    fill_only_fields = (
        "instansi",
        "alamat",
        "email",
        "nama_normalisasi",
        "whatsapp_normalized",
        "klasifikasi_produk",
    )
    for field in fill_only_fields:
        if not clean_text(existing[field]) and clean_text(imported.get(field)):
            updates[field] = imported[field]

    if not clean_text(existing["whatsapp"]) and imported.get("whatsapp_normalized"):
        updates["whatsapp"] = imported["whatsapp_normalized"]

    for field in (
        "nama_asli",
        "whatsapp_raw",
        "email_raw",
        "produk",
        "produk_existing",
        "sumber",
    ):
        merged = _merge_multi_value(existing[field], imported.get(field))
        if clean_text(merged) != clean_text(existing[field]):
            updates[field] = merged

    status = _merge_status(existing["status"], imported["status"])
    if status != existing["status"]:
        updates["status"] = status
    return updates


def analyze_database_duplicates(conn: sqlite3.Connection, consolidated: dict) -> dict:
    database_keys = _find_database_matches(conn)
    analyzed: list[dict] = []
    duplicate_database = 0
    for customer in consolidated["customers"]:
        match = None
        match_key = None
        for key in _identity_keys(customer):
            if key in database_keys:
                match = database_keys[key]
                match_key = key
                break

        result = dict(customer)
        if match is None:
            result.update({"action": "CREATE", "database_id": None, "database_match": None})
        else:
            duplicate_database += 1
            updates = _build_non_destructive_updates(match, customer)
            result.update(
                {
                    "action": "MERGE" if updates else "SKIP",
                    "database_id": match["id"],
                    "database_match": match_key[0] if match_key else None,
                    "updates": updates,
                    "before_values": {
                        field: match[field]
                        for field in updates
                    },
                }
            )
        analyzed.append(result)

    summary = dict(consolidated["summary"])
    summary.update(
        {
            "duplicate_database": duplicate_database,
            "create_count": sum(item["action"] == "CREATE" for item in analyzed),
            "merge_count": sum(item["action"] == "MERGE" for item in analyzed),
            "skip_count": sum(item["action"] == "SKIP" for item in analyzed),
        }
    )
    return {
        **consolidated,
        "customers": analyzed,
        "summary": summary,
    }


CUSTOMER_INSERT_FIELDS = (
    "nama",
    "whatsapp",
    "instansi",
    "produk",
    "sumber",
    "status",
    "nama_asli",
    "nama_normalisasi",
    "whatsapp_raw",
    "whatsapp_normalized",
    "email_raw",
    "email",
    "alamat",
    "produk_existing",
    "klasifikasi_produk",
    "import_batch_id",
)


def _insert_customer(conn: sqlite3.Connection, customer: dict, batch_id: str) -> int:
    values = dict(customer)
    values["import_batch_id"] = batch_id
    placeholders = ", ".join("?" for _ in CUSTOMER_INSERT_FIELDS)
    cursor = conn.execute(
        f"INSERT INTO customers ({', '.join(CUSTOMER_INSERT_FIELDS)}) VALUES ({placeholders})",
        tuple(values.get(field) for field in CUSTOMER_INSERT_FIELDS),
    )
    return cursor.lastrowid


def _update_customer(
    conn: sqlite3.Connection,
    customer_id: int,
    updates: dict,
) -> None:
    if not updates:
        return
    assignments = ", ".join(f"{field} = ?" for field in updates)
    conn.execute(
        f"UPDATE customers SET {assignments}, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
        (*updates.values(), customer_id),
    )


def import_customers_atomic(
    conn: sqlite3.Connection,
    consolidated: dict,
    *,
    batch_id: str,
    filename: str,
    file_sha256: str,
) -> dict:
    try:
        conn.execute("BEGIN IMMEDIATE")
        analyzed = analyze_database_duplicates(conn, consolidated)
        previous_batch = conn.execute(
            """
            SELECT id
            FROM customer_import_batches
            WHERE file_sha256 = ?
            ORDER BY created_at
            LIMIT 1
            """,
            (file_sha256,),
        ).fetchone()
        if previous_batch is not None:
            summary = {
                **analyzed["summary"],
                "created_count": 0,
                "merged_count": 0,
                "skipped_count": analyzed["summary"]["unique_ready"],
                "batch_id": batch_id,
                "duplicate_file": True,
                "previous_batch_id": previous_batch["id"],
            }
            report = build_import_report(
                summary,
                analyzed["per_sheet"],
                filename,
                file_sha256,
            )
            conn.commit()
            return {
                "summary": summary,
                "report_markdown": report,
                "customers": analyzed["customers"],
            }

        initial_summary = analyzed["summary"]
        conn.execute(
            """
            INSERT INTO customer_import_batches (
                id, filename, file_sha256, total_rows, unique_customers,
                warning_count, error_count
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                batch_id,
                filename,
                file_sha256,
                initial_summary["total_rows"],
                initial_summary["unique_ready"],
                initial_summary["warning_rows"],
                initial_summary["error_rows"],
            ),
        )

        created = merged = skipped = 0
        for customer in analyzed["customers"]:
            if customer["action"] == "CREATE":
                customer_id = _insert_customer(conn, customer, batch_id)
                changed_fields = [
                    field
                    for field in CUSTOMER_INSERT_FIELDS
                    if field != "import_batch_id" and customer.get(field) is not None
                ]
                before_values = None
                after_values = {
                    field: customer.get(field)
                    for field in changed_fields
                }
                created += 1
            elif customer["action"] == "MERGE":
                customer_id = customer["database_id"]
                changed_fields = list(customer["updates"])
                before_values = customer["before_values"]
                after_values = customer["updates"]
                _update_customer(conn, customer_id, customer["updates"])
                merged += 1
            else:
                customer_id = customer["database_id"]
                changed_fields = []
                before_values = None
                after_values = None
                skipped += 1

            customer["result_customer_id"] = customer_id
            customer["changed_fields"] = changed_fields
            conn.execute(
                """
                INSERT INTO customer_import_changes (
                    batch_id, customer_id, action, match_method, source_rows,
                    changed_fields, before_values, after_values
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    batch_id,
                    customer_id,
                    customer["action"],
                    customer.get("database_match"),
                    MULTI_VALUE_SEPARATOR.join(customer["source_rows"]),
                    MULTI_VALUE_SEPARATOR.join(changed_fields) or None,
                    json.dumps(before_values, ensure_ascii=False, sort_keys=True)
                    if before_values is not None
                    else None,
                    json.dumps(after_values, ensure_ascii=False, sort_keys=True)
                    if after_values is not None
                    else None,
                ),
            )

        summary = analyzed["summary"]
        result = {
            **summary,
            "created_count": created,
            "merged_count": merged,
            "skipped_count": skipped,
            "batch_id": batch_id,
        }
        report = build_import_report(
            result,
            analyzed["per_sheet"],
            filename,
            file_sha256,
            analyzed["customers"],
        )
        conn.execute(
            """
            UPDATE customer_import_batches
            SET created_count = ?,
                merged_count = ?,
                skipped_count = ?,
                report_markdown = ?,
                completed_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (
                created,
                merged,
                skipped,
                report,
                batch_id,
            ),
        )
        conn.commit()
        return {"summary": result, "report_markdown": report, "customers": analyzed["customers"]}
    except Exception:
        conn.rollback()
        raise


def build_import_report(
    summary: dict,
    per_sheet: dict,
    filename: str,
    file_sha256: str,
    customers: list[dict] | None = None,
) -> str:
    lines = [
        "# Laporan Import Master Customer",
        "",
        f"- File: `{filename}`",
        f"- SHA-256: `{file_sha256}`",
        f"- Batch: `{summary.get('batch_id', 'PREVIEW')}`",
        "",
        "## Rekap",
        "",
        f"- Total baris sumber: {summary['total_rows']}",
        f"- Customer unik: {summary['unique_ready']}",
        f"- Existing Customer: {summary['existing_customers']}",
        f"- Prospek: {summary['prospects']}",
        f"- Belum Terklasifikasi: {summary['unclassified']}",
        f"- Duplicate lintas sheet: {summary['duplicate_cross_sheet']}",
        f"- Duplicate database: {summary.get('duplicate_database', 0)}",
        f"- Warning row: {summary['warning_rows']}",
        f"- Error row: {summary['error_rows']}",
    ]
    if "created_count" in summary:
        lines.extend(
            (
                f"- Customer dibuat: {summary['created_count']}",
                f"- Customer di-merge: {summary['merged_count']}",
                f"- Customer dilewati: {summary['skipped_count']}",
            )
        )
    if summary.get("duplicate_file"):
        lines.append(
            f"- File identik sudah diproses pada batch: {summary['previous_batch_id']}"
        )
    lines.extend(("", "## Rekap per Sheet", "", "| Sheet | Baris | WA Valid | Warning | Error |", "|---|---:|---:|---:|---:|"))
    for sheet_name, values in per_sheet.items():
        lines.append(
            f"| {sheet_name} | {values['rows']} | {values['valid_phone']} | "
            f"{values['warning_rows']} | {values['error_rows']} |"
        )
    if customers:
        lines.extend(
            (
                "",
                "## Audit Action",
                "",
                "| Action | Customer ID | Match | Baris Sumber | Field Berubah |",
                "|---|---:|---|---|---|",
            )
        )
        for customer in customers:
            source_rows = ", ".join(customer.get("source_rows", []))
            changed_fields = ", ".join(customer.get("changed_fields", [])) or "-"
            lines.append(
                f"| {customer.get('action', '-')} | "
                f"{customer.get('result_customer_id') or customer.get('database_id') or '-'} | "
                f"{customer.get('database_match') or '-'} | {source_rows} | {changed_fields} |"
            )
    return "\n".join(lines) + "\n"


def sha256_bytes(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def prepare_customer_import(content: bytes, conn: sqlite3.Connection) -> dict:
    parsed = parse_customer_workbook(content)
    consolidated = consolidate_customer_records(parsed)
    return analyze_database_duplicates(conn, consolidated)
