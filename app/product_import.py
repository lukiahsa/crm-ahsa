"""Parser dan import atomik untuk master produk resmi berbentuk XLSX."""

import re
import unicodedata
from copy import deepcopy
from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import load_workbook


MAX_IMPORT_FILE_SIZE = 5 * 1024 * 1024
MAX_IMPORT_ROWS = 5000
KNOWN_CATEGORIES = {
    "Tempat Sampah",
    "Tangga",
    "Material Handling",
}


class ProductImportError(ValueError):
    """Kesalahan file atau struktur import yang aman ditampilkan ke user."""


def normalize_text(value):
    """Normalisasi ringan tanpa mengubah arti atau kode sumber."""
    if value is None:
        return None

    if isinstance(value, bool):
        text = str(value)
    elif isinstance(value, int):
        text = str(value)
    elif isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value)

    normalized = re.sub(r"\s+", " ", text).strip()
    return normalized or None


def parse_rupiah_integer(value):
    """Ubah nilai Rupiah menjadi integer tanpa perhitungan float."""
    if value is None or normalize_text(value) is None:
        return None
    if isinstance(value, bool):
        raise ProductImportError("Harga harus berupa angka Rupiah.")

    if isinstance(value, int):
        return value

    raw = normalize_text(value)
    cleaned = re.sub(r"(?i)^rp\.?\s*", "", raw).replace(" ", "")

    if re.fullmatch(r"[-+]?\d{1,3}(\.\d{3})+", cleaned):
        cleaned = cleaned.replace(".", "")
    elif re.fullmatch(r"[-+]?\d{1,3}(,\d{3})+", cleaned):
        cleaned = cleaned.replace(",", "")
    elif "," in cleaned and "." not in cleaned:
        cleaned = cleaned.replace(",", ".")

    try:
        amount = Decimal(cleaned)
    except InvalidOperation as error:
        raise ProductImportError("Harga harus berupa angka Rupiah.") from error

    if not amount.is_finite() or amount != amount.to_integral_value():
        raise ProductImportError("Harga Rupiah harus berupa bilangan bulat.")

    return int(amount)


def _slug(value):
    normalized = normalize_text(value) or ""
    ascii_text = unicodedata.normalize("NFKD", normalized).encode(
        "ascii", "ignore"
    ).decode("ascii")
    return re.sub(r"[^A-Za-z0-9]+", "-", ascii_text).strip("-").upper()


def build_trash_sku(capacity, color):
    """Buat SKU Tempat Sampah yang stabil dari kapasitas/tipe dan warna."""
    capacity_text = normalize_text(capacity) or ""
    capacity_text = re.sub(
        r"(?i)^tempat\s+sampah\s*", "", capacity_text
    ).strip()

    match = re.match(r"^(\d+)\s+Liter(?:\s+(.*))?$", capacity_text, re.I)
    parts = ["TS"]
    if match:
        parts.append(match.group(1).zfill(3))
        descriptor = _slug(match.group(2)) if match.group(2) else ""
        if descriptor:
            parts.append(descriptor)
    else:
        capacity_slug = _slug(capacity_text)
        if capacity_slug:
            parts.append(capacity_slug)

    color_slug = _slug(color)
    if color_slug:
        parts.append(color_slug)

    return "-".join(parts)


def _base_row(source_group, source_row):
    return {
        "source_group": source_group,
        "source_row": source_row,
        "kategori": source_group,
        "subkategori": None,
        "nama_produk": None,
        "kode_produk": None,
        "sku": None,
        "brand": None,
        "variant": None,
        "warna": None,
        "ukuran": None,
        "jenis_produk": None,
        "steps": None,
        "satuan": "Unit",
        "harga_modal_default": None,
        "warnings": [],
        "errors": [],
    }


def _validate_row(row):
    if not row["nama_produk"]:
        row["errors"].append("Nama produk wajib diisi.")
    if row["kategori"] not in KNOWN_CATEGORIES:
        row["errors"].append("Kategori produk tidak dikenal.")

    amount = row["harga_modal_default"]
    if amount is None:
        row["warnings"].append(
            "Harga modal kosong; disimpan 0 sesuai schema existing."
        )
    elif amount < 0:
        row["errors"].append("Harga modal tidak boleh negatif.")

    return row


def _parse_price(value, row):
    try:
        row["harga_modal_default"] = parse_rupiah_integer(value)
    except ProductImportError as error:
        row["errors"].append(str(error))


def _parse_trash_rows(source_rows):
    rows = []
    last_name = None
    last_capacity = None

    for source_row, values in source_rows:
        raw_name, raw_capacity, raw_color, raw_price = values[1:5]
        if all(
            normalize_text(value) is None
            for value in (raw_name, raw_capacity, raw_color, raw_price)
        ):
            continue

        last_name = normalize_text(raw_name) or last_name
        last_capacity = normalize_text(raw_capacity) or last_capacity
        color = normalize_text(raw_color)

        row = _base_row("Tempat Sampah", source_row)
        row["variant"] = last_capacity
        row["warna"] = color
        row["brand"] = "Dalton"

        if last_name and last_capacity:
            if last_capacity.casefold().startswith(last_name.casefold()):
                name = last_capacity
            else:
                name = f"{last_name} {last_capacity}"
            row["nama_produk"] = f"{name} - {color}" if color else name
            row["kode_produk"] = build_trash_sku(last_capacity, color)
            row["sku"] = row["kode_produk"]

        _parse_price(raw_price, row)
        rows.append(_validate_row(row))

    return rows


def _parse_ladder_rows(source_rows):
    rows = []
    last_name = None
    last_brand = None
    last_product_type = None

    for source_row, source_values in source_rows:
        values = source_values[6:13]
        if all(normalize_text(value) is None for value in values):
            continue

        raw_name, raw_brand, raw_product_type, raw_code, raw_size, raw_steps, raw_price = values
        last_name = normalize_text(raw_name) or last_name
        last_brand = normalize_text(raw_brand) or last_brand
        last_product_type = normalize_text(raw_product_type) or last_product_type

        code = normalize_text(raw_code)
        row = _base_row("Tangga", source_row)
        row.update(
            {
                "brand": last_brand,
                "jenis_produk": last_product_type,
                "kode_produk": code,
                "sku": code,
                "ukuran": normalize_text(raw_size),
                "steps": normalize_text(raw_steps),
            }
        )

        name_parts = [last_name, last_product_type, code]
        row["nama_produk"] = " ".join(part for part in name_parts if part)
        if not code:
            row["errors"].append("Tipe/kode produk Tangga wajib diisi.")

        _parse_price(raw_price, row)
        rows.append(_validate_row(row))

    return rows


def _parse_material_rows(source_rows):
    rows = []
    last_subcategory = None

    for source_row, values in source_rows:
        raw_name, raw_code, raw_price = values[14:17]
        if all(
            normalize_text(value) is None
            for value in (raw_name, raw_code, raw_price)
        ):
            continue

        last_subcategory = normalize_text(raw_name) or last_subcategory
        code = normalize_text(raw_code)
        row = _base_row("Material Handling", source_row)
        row["subkategori"] = last_subcategory
        row["kode_produk"] = code
        row["sku"] = code
        row["nama_produk"] = " ".join(
            part for part in (last_subcategory, code) if part
        )
        if not code:
            row["warnings"].append(
                "Tipe/kode produk kosong; duplicate dicek dari identitas produk."
            )

        _parse_price(raw_price, row)
        rows.append(_validate_row(row))

    return rows


def parse_product_workbook(file_bytes):
    """Parse tiga kelompok pada workbook resmi menjadi baris ternormalisasi."""
    if not file_bytes:
        raise ProductImportError("File XLSX wajib dipilih.")
    if len(file_bytes) > MAX_IMPORT_FILE_SIZE:
        raise ProductImportError("Ukuran file melebihi batas 5 MB.")

    try:
        workbook = load_workbook(
            BytesIO(file_bytes), data_only=True, read_only=True, keep_links=False
        )
    except Exception as error:
        raise ProductImportError("File XLSX tidak dapat dibaca.") from error

    try:
        sheet = workbook.active
        workbook_rows = []
        for source_row, values in enumerate(
            sheet.iter_rows(min_row=3, max_col=17, values_only=True),
            start=3,
        ):
            if source_row > MAX_IMPORT_ROWS + 4:
                raise ProductImportError(
                    f"Workbook melebihi batas {MAX_IMPORT_ROWS} baris data."
                )
            workbook_rows.append((source_row, values))
        if len(workbook_rows) < 2:
            raise ProductImportError("Workbook tidak memiliki header produk.")

        group_header = workbook_rows[0][1]
        header = workbook_rows[1][1]
        source_rows = workbook_rows[2:]
        expected_groups = {
            1: ("B3", "Tempat Sampah"),
            6: ("G3", "Tangga"),
            14: ("O3", "Material Handling"),
        }
        expected_headers = {
            1: ("B4", "Nama Produk"),
            2: ("C4", "Kapasitas"),
            3: ("D4", "Warna"),
            4: ("E4", "Harga Modal"),
            6: ("G4", "Nama Produk"),
            7: ("H4", "Merk"),
            8: ("I4", "Jenis Produk"),
            9: ("J4", "Tipe Produk"),
            10: ("K4", "Ukuran"),
            11: ("L4", "Steps"),
            12: ("M4", "Harga Modal"),
            14: ("O4", "Nama Produk"),
            15: ("P4", "Tipe"),
            16: ("Q4", "Harga Modal"),
        }
        invalid = [
            cell
            for index, (cell, expected) in expected_groups.items()
            if normalize_text(group_header[index]) != expected
        ] + [
            cell
            for index, (cell, expected) in expected_headers.items()
            if normalize_text(header[index]) != expected
        ]
        if invalid:
            raise ProductImportError(
                "Struktur workbook tidak sesuai template MASTER PRODUK.xlsx "
                f"pada kolom: {', '.join(invalid)}."
            )

        return (
            _parse_trash_rows(source_rows)
            + _parse_ladder_rows(source_rows)
            + _parse_material_rows(source_rows)
        )
    finally:
        workbook.close()


def _key(value):
    return (normalize_text(value) or "").casefold()


def _row_value(row, field):
    if isinstance(row, dict):
        return row.get(field)
    return row[field]


def _composite_key(row):
    return (
        _key(_row_value(row, "nama_produk")),
        _key(_row_value(row, "brand")),
        _key(_row_value(row, "variant")),
        _key(_row_value(row, "ukuran")),
        _key(_row_value(row, "warna")),
    )


def analyze_product_rows(conn, rows):
    """Klasifikasikan valid/warning/duplicate/error secara deterministik."""
    analyzed = deepcopy(rows)
    existing_rows = conn.execute(
        """
        SELECT
            products.id,
            products.kode_produk,
            products.nama_produk,
            product_brands.nama AS brand,
            product_variants.nama AS variant,
            product_sizes.nama AS ukuran,
            product_colors.nama AS warna
        FROM products
        LEFT JOIN product_brands
            ON products.brand_id = product_brands.id
        LEFT JOIN product_variants
            ON products.variant_id = product_variants.id
        LEFT JOIN product_sizes
            ON products.size_id = product_sizes.id
        LEFT JOIN product_colors
            ON products.color_id = product_colors.id
        """
    ).fetchall()

    existing_codes = {
        _key(row["kode_produk"]): row
        for row in existing_rows
        if _key(row["kode_produk"])
    }
    existing_composites = {
        _composite_key(row): row for row in existing_rows
    }
    seen_codes = {}
    seen_composites = {}

    for row in analyzed:
        row["duplicate_kind"] = None
        row["duplicate_message"] = None
        if row["errors"]:
            row["status"] = "error"
            continue

        code_key = _key(row["kode_produk"])
        composite = _composite_key(row)
        duplicate_kind = None
        duplicate_message = None

        if code_key and code_key in existing_codes:
            existing = existing_codes[code_key]
            if _key(existing["nama_produk"]) != _key(row["nama_produk"]):
                duplicate_kind = "code_name_conflict"
                duplicate_message = (
                    "Kode produk sudah dipakai database untuk nama berbeda."
                )
            else:
                duplicate_kind = "existing_database"
                duplicate_message = "Produk sudah ada di database."
        elif code_key and code_key in seen_codes:
            first = seen_codes[code_key]
            if _key(first["nama_produk"]) != _key(row["nama_produk"]):
                duplicate_kind = "code_name_conflict"
                duplicate_message = (
                    "Kode produk dipakai baris lain untuk nama berbeda."
                )
            else:
                duplicate_kind = "duplicate_in_file"
                duplicate_message = "Produk berulang di dalam file."
        elif composite in existing_composites:
            duplicate_kind = "existing_database"
            duplicate_message = (
                "Kombinasi nama, brand, varian/ukuran, dan warna sudah ada."
            )
        elif composite in seen_composites:
            duplicate_kind = "duplicate_in_file"
            duplicate_message = (
                "Kombinasi nama, brand, varian/ukuran, dan warna berulang."
            )

        if duplicate_kind:
            row["status"] = "duplicate"
            row["duplicate_kind"] = duplicate_kind
            row["duplicate_message"] = duplicate_message
            continue

        if code_key:
            seen_codes[code_key] = row
        seen_composites[composite] = row
        row["status"] = "warning" if row["warnings"] else "valid"

    return analyzed


def summarize_rows(rows):
    summary = {
        "total": len(rows),
        "valid": 0,
        "warning": 0,
        "duplicate": 0,
        "error": 0,
    }
    for row in rows:
        status = row.get("status", "error")
        summary[status] = summary.get(status, 0) + 1
    return summary


def _get_or_create_simple_reference(conn, table_name, name):
    allowed = {
        "product_categories",
        "product_brands",
        "product_colors",
    }
    if table_name not in allowed or not name:
        return None
    existing = conn.execute(
        f"SELECT id FROM {table_name} WHERE LOWER(nama) = LOWER(?)",
        (name,),
    ).fetchone()
    if existing:
        return existing["id"]
    return conn.execute(
        f"INSERT INTO {table_name} (nama) VALUES (?)", (name,)
    ).lastrowid


def _get_or_create_category_reference(conn, table_name, category_id, name):
    if table_name not in {"product_variants", "product_sizes"} or not name:
        return None
    existing = conn.execute(
        f"""
        SELECT id FROM {table_name}
        WHERE LOWER(nama) = LOWER(?)
          AND ((category_id = ?) OR (category_id IS NULL AND ? IS NULL))
        """,
        (name, category_id, category_id),
    ).fetchone()
    if existing:
        return existing["id"]
    return conn.execute(
        f"INSERT INTO {table_name} (category_id, nama) VALUES (?, ?)",
        (category_id, name),
    ).lastrowid


def _insert_product(conn, row):
    category_id = _get_or_create_simple_reference(
        conn, "product_categories", row["kategori"]
    )
    brand_id = _get_or_create_simple_reference(
        conn, "product_brands", row["brand"]
    )
    variant_id = _get_or_create_category_reference(
        conn, "product_variants", category_id, row["variant"]
    )
    color_id = _get_or_create_simple_reference(
        conn, "product_colors", row["warna"]
    )
    size_id = _get_or_create_category_reference(
        conn, "product_sizes", category_id, row["ukuran"]
    )

    conn.execute(
        """
        INSERT INTO products (
            kode_produk,
            nama_produk,
            category_id,
            brand_id,
            variant_id,
            color_id,
            size_id,
            satuan,
            harga_jual_default,
            harga_modal_default,
            status_aktif,
            subkategori,
            jenis_produk,
            steps
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, 0, ?, 1, ?, ?, ?)
        """,
        (
            row["kode_produk"],
            row["nama_produk"],
            category_id,
            brand_id,
            variant_id,
            color_id,
            size_id,
            row["satuan"],
            row["harga_modal_default"] or 0,
            row["subkategori"],
            row["jenis_produk"],
            row["steps"],
        ),
    )


def import_product_rows(conn, rows):
    """Re-check duplicate lalu import semua baris dalam satu transaction."""
    try:
        conn.execute("BEGIN IMMEDIATE")
        analyzed = analyze_product_rows(conn, rows)
        created = 0
        for row in analyzed:
            if row["status"] not in {"valid", "warning"}:
                continue
            _insert_product(conn, row)
            created += 1
        conn.commit()
    except Exception:
        conn.rollback()
        raise

    summary = summarize_rows(analyzed)
    summary.update(
        {
            "created": created,
            "skipped": summary["duplicate"] + summary["error"],
        }
    )
    return analyzed, summary
